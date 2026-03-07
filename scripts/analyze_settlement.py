#!/usr/bin/env python3
"""分析结算单数据，提取MSKU单价分配规律"""
import os
import json
from openpyxl import load_workbook
from collections import defaultdict

SETTLEMENT_DIR = "docs/FBA货单信息/3、结算单"

# 站点序号映射
SITE_MAP = {"001": "USD", "002": "CAD", "003": "GBP", "004": "EUR"}

results = []

for filename in sorted(os.listdir(SETTLEMENT_DIR)):
    if not filename.endswith(".xlsx") or filename.startswith("~"):
        continue
    
    filepath = os.path.join(SETTLEMENT_DIR, filename)
    # 从文件名提取编号和站点
    doc_no = filename.split("-")[0]
    site_suffix = doc_no[-3:]  # 001/002/003/004
    site = SITE_MAP.get(site_suffix, "UNKNOWN")
    period_date = doc_no[:8]
    
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # 读取结算周期 (B5)
        period = ws["B5"].value if ws["B5"].value else ""
        
        # 读取明细行 (从第9行开始，到"合计Total"行结束)
        items = []
        for row in range(9, ws.max_row + 1):
            no_val = ws.cell(row=row, column=1).value
            desc = ws.cell(row=row, column=2).value
            currency = ws.cell(row=row, column=3).value
            unit_price = ws.cell(row=row, column=4).value
            qty = ws.cell(row=row, column=5).value
            amount = ws.cell(row=row, column=6).value
            
            # 跳过合计行和空行
            if desc and "Total" in str(desc):
                break
            if desc and "合计" in str(desc):
                break
            if not desc or not qty:
                continue
            
            items.append({
                "msku": str(desc).strip(),
                "currency": str(currency).strip() if currency else site,
                "unit_price": float(unit_price) if unit_price else 0,
                "quantity": int(qty) if qty else 0,
                "amount": float(amount) if amount else 0
            })
        
        if items:
            results.append({
                "doc_no": doc_no,
                "period_date": period_date,
                "site": site,
                "period": str(period).strip() if period else "",
                "items": items
            })
        
        wb.close()
    except Exception as e:
        print(f"Error reading {filename}: {e}")

# 分析1: 每个站点有哪些MSKU
print("=" * 80)
print("分析1: 各站点MSKU列表")
print("=" * 80)
site_mskus = defaultdict(set)
for r in results:
    for item in r["items"]:
        site_mskus[r["site"]].add(item["msku"])

for site in ["USD", "CAD", "GBP", "EUR"]:
    mskus = sorted(site_mskus.get(site, []))
    print(f"\n{site} 站点 ({len(mskus)} 个MSKU):")
    for m in mskus:
        print(f"  {m}")

# 分析2: 同一站点同一周期内，不同MSKU的单价是否相同
print("\n" + "=" * 80)
print("分析2: 同站点同周期内MSKU单价对比")
print("=" * 80)
diff_count = 0
same_count = 0
for r in results:
    prices = set()
    for item in r["items"]:
        prices.add(round(item["unit_price"], 4))
    if len(prices) > 1:
        diff_count += 1
        print(f"\n{r['doc_no']} ({r['site']}) - 单价不同!")
        for item in r["items"]:
            print(f"  {item['msku']}: {item['unit_price']:.4f} x {item['quantity']} = {item['amount']:.4f}")
    else:
        same_count += 1

print(f"\n统计: 单价相同={same_count}, 单价不同={diff_count}")

# 分析3: MSKU特征分析 - 解析MSKU结构
print("\n" + "=" * 80)
print("分析3: MSKU结构解析")
print("=" * 80)
all_mskus = set()
for r in results:
    for item in r["items"]:
        all_mskus.add(item["msku"])

for msku in sorted(all_mskus):
    parts = msku.split("-")
    prefix = parts[0] if len(parts) > 0 else ""
    model = parts[1] if len(parts) > 1 else ""
    variant = parts[2] if len(parts) > 2 else ""
    spec = parts[3] if len(parts) > 3 else ""
    print(f"  {msku:30s} → 前缀={prefix}, 型号={model}, 变体={variant}, 规格={spec}")

# 分析4: 跨周期单价变化趋势（按MSKU）
print("\n" + "=" * 80)
print("分析4: 各MSKU跨周期单价变化")
print("=" * 80)
msku_prices = defaultdict(list)
for r in results:
    for item in r["items"]:
        msku_prices[item["msku"]].append({
            "date": r["period_date"],
            "price": item["unit_price"],
            "qty": item["quantity"],
            "site": r["site"]
        })

for msku in sorted(msku_prices.keys()):
    records = msku_prices[msku]
    prices = [r["price"] for r in records]
    min_p = min(prices)
    max_p = max(prices)
    avg_p = sum(prices) / len(prices)
    print(f"\n{msku} ({records[0]['site']}):")
    print(f"  周期数={len(records)}, 最低={min_p:.4f}, 最高={max_p:.4f}, 平均={avg_p:.4f}, 波动={max_p-min_p:.4f}")
    # 显示前5个和后5个周期
    for rec in records[:3]:
        print(f"    {rec['date']}: 单价={rec['price']:.4f}, 数量={rec['qty']}")
    if len(records) > 6:
        print(f"    ... (省略 {len(records)-6} 个周期)")
    for rec in records[-3:]:
        print(f"    {rec['date']}: 单价={rec['price']:.4f}, 数量={rec['qty']}")

# 分析5: 同一站点同一周期内，不同MSKU单价的比例关系
print("\n" + "=" * 80)
print("分析5: 同站点同周期内MSKU单价比例关系（仅单价不同的周期）")
print("=" * 80)
for r in results:
    prices = {}
    for item in r["items"]:
        prices[item["msku"]] = item["unit_price"]
    unique_prices = set(round(p, 4) for p in prices.values())
    if len(unique_prices) > 1:
        # 找最低单价作为基准
        base_price = min(prices.values())
        print(f"\n{r['doc_no']} ({r['site']}, {r['period']}):")
        for msku, price in sorted(prices.items()):
            ratio = price / base_price if base_price > 0 else 0
            qty = next(i["quantity"] for i in r["items"] if i["msku"] == msku)
            print(f"  {msku:30s}: 单价={price:.4f}, 数量={qty}, 比例={ratio:.4f}")

# 输出JSON供后续分析
with open("scripts/settlement_analysis.json", "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=2)

print("\n\n分析完成，详细数据已保存到 scripts/settlement_analysis.json")
